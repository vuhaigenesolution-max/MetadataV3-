"""
Checksamplenumber.py
--------------------
Đối soát file SUM (multi-page Pool Allocation) ↔ folder metadata.
Trả về các sample có trong meta nhưng KHÔNG có trong SUM.

Public API (giữ nguyên cho backend):
  - read_meta_folder(folder)        → df_meta  (RunName + 4 cột Excel A,B,U,V từ row 21+)
  - process_sum_file(sum_path)      → df_sum   (RunName, Col 3, Col 5, sample_order, Loại, Sample Type)
  - find_meta_only(df_meta, df_sum) → df_meta_only (RunName, expNum, sampleOrder [+ debug keys])

Layout file SUM:
  - Multi-page lặp ngang trên 1 sheet
  - Mỗi page = 20 cột data + 2 cột trống (PAGE_STEP=22)
  - Row 1 mỗi page: A1=RunContent, E1=RunNameNotes, Q1=MachineNotes
  - Row 4 (SUM_HEADER_ROW): header (sẽ thay bằng Col 1..Col 20)
  - Row 5+ (SUM_DATA_ROW): data, trim theo Col 3 (cột C trong page)

Layout file metadata:
  - Filename: metadata_<RunName>_<yyyymmdd>.xlsx (RunName = R/P + 4-5 số)
  - Sheet "Sample"
  - Row 21: header — lấy 4 cột Excel A, B, U, V (vị trí, KHÔNG phải tên cột)
  - Row 22+: data

Match rule (mỗi row meta cần khớp 1 trong 2 đường):
  - NIPT  : RunName + expNum (col A) + sampleOrder (col B)
            với SUM Sample Type='NIPT' AND Loại=''
  - other : RunName + (V hoặc fallback U)
            với SUM Sample Type='other' (KHÔNG cần sample_order)
"""

import os
import re
import glob
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string


# ══════════════════════════════════════════════════════════
# CONSTANTS
# ══════════════════════════════════════════════════════════

# SUM page layout
PAGE_COLS      = 20    # số cột mỗi page (A..T)
PAGE_STEP      = 22    # 20 cột data + 2 cột trống
SUM_HEADER_ROW = 4     # header trong page (Excel row, 1-based)
SUM_DATA_ROW   = 5     # data trong page bắt đầu từ row này
SUM_KEY_COL    = "C"   # cột key trong page để trim → tương đương Col 3

# Metadata file layout
META_HEADER_ROW  = 21
META_DATA_ROW    = 22
META_TARGET_COLS = [0, 1, 20, 21]   # 0-indexed: A, B, U, V

# Patterns
META_FILE_PATTERN = re.compile(r"metadata_([RP]\d{4,5})_(\d{8})", re.IGNORECASE)
RANGE_PATTERN     = re.compile(r"^\s*(\d+)\s*-\s*(\d+)\s*$")
# NIPT prefixes (cần khớp sample_order): EK, TK, GS hoặc 1 chữ trong [ETBHLVP] + digit
NIPT_PATTERN      = re.compile(r"^(?:EK|TK|GS|[ETBHLVP])\d+", flags=re.IGNORECASE)
# CRH = pool NIPT: CR / CRH + digit (KHÔNG cần sample_order, vẫn match với col A bên meta)
CRH_PATTERN       = re.compile(r"^(?:CRH|CR)\d+", flags=re.IGNORECASE)


# ══════════════════════════════════════════════════════════
# 1. ĐỌC FILE SUM
# ══════════════════════════════════════════════════════════

def _is_blank(x) -> bool:
    """Coi là rỗng nếu None hoặc string toàn whitespace."""
    if x is None:
        return True
    if isinstance(x, str) and x.strip() == "":
        return True
    return False


def _detect_page_starts(ws) -> list:
    """Tự dò danh sách col_start của các page có data.
    Quét từng block PAGE_STEP cột; page hợp lệ nếu 4 row đầu có ít nhất 1 cell có data.
    Dừng khi gặp block hoàn toàn rỗng.
    """
    starts = []
    max_col = ws.max_column
    probe_rows = min(ws.max_row, 4)

    for col_start in range(1, max_col + 1, PAGE_STEP):
        col_end = min(col_start + PAGE_COLS - 1, max_col)
        has_data = any(
            not _is_blank(ws.cell(r, c).value)
            for r in range(1, probe_rows + 1)
            for c in range(col_start, col_end + 1)
        )
        if has_data:
            starts.append(col_start)
        else:
            break
    return starts


def _read_one_page(ws, col_start: int) -> pd.DataFrame:
    """Đọc 1 page (PAGE_COLS cột bắt đầu từ col_start).
    - Trim hàng cuối theo cột key (Col 3 trong page = Excel cột C).
    - Expand merged cells (lan giá trị top-left ra cả range).
    Trả về df 20 cột raw (chưa cắt header).
    """
    col_end    = col_start + PAGE_COLS - 1
    max_row    = ws.max_row
    key_offset = column_index_from_string(SUM_KEY_COL) - 1   # C → 2

    # 1. Đọc raw 20 cột
    data = [
        [ws.cell(r, c).value for c in range(col_start, col_end + 1)]
        for r in range(1, max_row + 1)
    ]

    # 2. Trim theo key column (TRƯỚC khi expand merge)
    df_raw = pd.DataFrame(data).replace(r'^\s*$', np.nan, regex=True)
    start_df_row = SUM_DATA_ROW - 1
    if df_raw.shape[0] > start_df_row:
        s = df_raw.iloc[start_df_row:, key_offset]
        m = s.notna() & s.astype(str).str.strip().ne("")
        if m.any():
            last_idx = start_df_row + np.where(m.to_numpy())[0].max()
            data = data[: last_idx + 1]
        else:
            data = data[: max(SUM_HEADER_ROW, start_df_row)]

    # 3. Expand merged cells (giữ giá trị top-left cho cả range)
    n_rows = len(data)
    for mr in ws.merged_cells.ranges:
        c1 = max(mr.min_col, col_start)
        c2 = min(mr.max_col, col_end)
        if c1 > c2:
            continue
        r1 = max(mr.min_row, 1)
        r2 = min(mr.max_row, n_rows)
        if r1 > r2:
            continue
        top_left = ws.cell(mr.min_row, mr.min_col).value
        if top_left is None:
            continue
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                data[r - 1][c - col_start] = top_left

    return pd.DataFrame(data).replace(r'^\s*$', np.nan, regex=True)


def parse_file_by_pages(source_file_path: str, sheet_name: str):
    """Duyệt mọi page → 1 DataFrame phẳng + 3 cột meta (RunContent, RunNameNotes, MachineNotes).
    Return: (df, page_count)
    """
    wb = load_workbook(source_file_path, data_only=True)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]

    page_starts = _detect_page_starts(ws)
    df_pages = []

    for col_start in page_starts:
        df_page = _read_one_page(ws, col_start)

        # 3 ô meta ở row 1 (idx 0): A1, E1, Q1
        run_content   = df_page.iat[0, 0]    # A1
        run_name_note = df_page.iat[0, 4]    # E1
        machine_note  = df_page.iat[0, 16]   # Q1

        # Skip 3 dòng đầu, dòng 4 (idx 3) làm header rồi bỏ
        body = df_page.iloc[3:].copy()
        body.columns = body.iloc[0].astype(str)
        body = body.iloc[1:].reset_index(drop=True)

        # Ép về 20 cột với tên Col 1..Col 20
        body = body.iloc[:, :PAGE_COLS]
        body.columns = [f"Col {j}" for j in range(1, PAGE_COLS + 1)]

        # Trim đến row đầu Col 3 trống
        c3_str = body["Col 3"].astype(str).str.strip()
        if c3_str.eq("").any():
            body = body.iloc[:c3_str.eq("").idxmax()].copy()

        # Giữ row có cả Col 1 lẫn Col 3 không trống
        keep = (
            body["Col 1"].notna() & body["Col 1"].astype(str).str.strip().ne("") &
            body["Col 3"].notna() & body["Col 3"].astype(str).str.strip().ne("")
        )
        body = body.loc[keep].reset_index(drop=True)

        body["RunContent"]   = run_content
        body["RunNameNotes"] = run_name_note
        body["MachineNotes"] = machine_note
        df_pages.append(body)

    wb.close()
    df_all = pd.concat(df_pages, ignore_index=True) if df_pages else pd.DataFrame()
    return df_all, len(page_starts)


def source_to_df(source_path: str, sheet: str) -> pd.DataFrame:
    """Đọc file SUM, trích RunName từ RunContent (lấy phần sau ':' đến ',')."""
    df, _ = parse_file_by_pages(source_path, sheet)
    df = df.copy()
    df["RunName"] = (
        df["RunContent"].astype(str)
        .str.extract(r":\s*([^,]+)", expand=False)
        .str.strip()
    )
    return df


# ══════════════════════════════════════════════════════════
# 2. PARSE Col 5 (range) + CLASSIFY Col 3 (NIPT / other)
# ══════════════════════════════════════════════════════════

def _parse_col5(val):
    """Col 5 → (sample_orders[], loai_flags[]).
       NaN / ''     → 1 dòng sample_order='', loai='' (giữ row, sample_order trống)
       'X-Y'        → range X..Y, loai='' cho mọi số
       'X-Y/<excl>' → range X..Y, loai='Loại' cho số trong <excl>
       khác         → 1 dòng giữ nguyên giá trị, loai=''
    """
    # NaN / None / empty
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return [""], [""]
    s = str(val).strip()
    if s == "" or s.lower() == "nan":
        return [""], [""]

    main_part, _, exc_part = s.partition("/")

    m = RANGE_PATTERN.match(main_part.strip())
    if not m:
        return [s], [""]

    start, end = int(m.group(1)), int(m.group(2))
    if end < start:
        return [s], [""]

    orders = list(range(start, end + 1))

    # Parse exclusion list (nếu có)
    excluded = set()
    for token in exc_part.split(","):
        token = token.strip()
        if not token:
            continue
        m2 = RANGE_PATTERN.match(token)
        if m2:
            a, b = int(m2.group(1)), int(m2.group(2))
            excluded.update(range(a, b + 1))
        else:
            try:
                excluded.add(int(token))
            except ValueError:
                pass

    loai = ["Loại" if o in excluded else "" for o in orders]
    return orders, loai


def _classify_col3(val) -> str:
    """Col 3 → 'NIPT' | 'CRH' | 'other'.

    NIPT (match với col A meta, CẦN khớp sample_order):
      - PGS, PGTM (exact)
      - EK<digits>, TK<digits>, GS<digits>
      - E/T/B/H/L/V/P + <digits>

    CRH (pool NIPT — match với col A meta, KHÔNG cần sample_order):
      - CR<digits>, CRH<digits>  (đi theo nguyên pool, chạy 1 lần)

    other (match với col V/U meta, KHÔNG cần sample_order):
      - tất cả các mã còn lại
    """
    s = str(val).strip().upper()
    if s in ("PGS", "PGTM"):
        return "NIPT"
    if CRH_PATTERN.match(s):
        return "CRH"
    if NIPT_PATTERN.match(s):
        return "NIPT"
    return "other"


def process_sum_file(sum_path: str, sheet: str = "Sheet1") -> pd.DataFrame:
    """Pipeline đầy đủ cho file SUM.
    Output 6 cột: RunName, Col 3, Col 5, sample_order, Loại, Sample Type.

    - Col 3 đã được clean: bỏ phần "(...)" (vd ABC(XYZ) → ABC).
    - Col 5 đã được explode thành nhiều dòng (theo range).
    - Row chỉ bị drop khi RunName HOẶC Col 3 trống.
      Col 5 trống vẫn giữ row (sample_order='', Loại='').
    """
    df = source_to_df(sum_path, sheet)

    df_out = (
        df[["RunName", "Col 3", "Col 5"]]
        .replace(r"^\s*$", np.nan, regex=True)
        .dropna(subset=["RunName", "Col 3"])     # KHÔNG drop khi Col 5 trống
        .reset_index(drop=True)
    )

    # Clean Col 3: bỏ "(...)"
    df_out["Col 3"] = (
        df_out["Col 3"].astype(str)
        .str.replace(r"\s*\([^)]*\)", "", regex=True)
        .str.strip()
    )

    # Explode Col 5 → sample_order + Loại
    parsed = df_out["Col 5"].apply(_parse_col5)
    df_out["sample_order"] = parsed.apply(lambda x: x[0])
    df_out["Loại"]         = parsed.apply(lambda x: x[1])
    df_out = df_out.explode(["sample_order", "Loại"]).reset_index(drop=True)

    # Classify NIPT vs other
    df_out["Sample Type"] = df_out["Col 3"].apply(_classify_col3)
    return df_out


# ══════════════════════════════════════════════════════════
# 3. ĐỌC FOLDER METADATA
# ══════════════════════════════════════════════════════════

def read_meta_folder(folder: str, sheet_name: str = "Sample") -> pd.DataFrame:
    """Đọc tất cả file metadata trong folder → 1 DataFrame gộp.

    Mỗi file:
      - RunName : regex 'metadata_([RP]\\d{4,5})_(\\d{8})' từ tên file
                   (fallback: stem nếu không match)
      - Header  : Excel row 21, lấy 4 cột Excel A, B, U, V
      - Data    : Excel row 22+, cùng 4 cột

    Lưu ý: A/B/U/V là CỘT EXCEL (vị trí), KHÔNG phải tên cột.

    Return: DataFrame [RunName, header_A, header_B, header_U, header_V].
    """
    files = sorted(
        glob.glob(os.path.join(folder, "*.xlsx"))
        + glob.glob(os.path.join(folder, "*.xls"))
    )

    df_lst = []
    for fp in files:
        fname = os.path.basename(fp)
        m = META_FILE_PATTERN.search(fname)
        runname = m.group(1) if m else os.path.splitext(fname)[0]

        wb = load_workbook(fp, data_only=True, read_only=True)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]

        # Header row 21
        header_row = next(
            ws.iter_rows(min_row=META_HEADER_ROW, max_row=META_HEADER_ROW, values_only=True),
            (),
        )
        headers = [
            str(header_row[i]).strip()
            if i < len(header_row) and header_row[i] not in (None, "")
            else f"Col_{i + 1}"
            for i in META_TARGET_COLS
        ]

        # Data từ row 22+
        rows = []
        for row in ws.iter_rows(min_row=META_DATA_ROW, values_only=True):
            picked = [row[i] if i < len(row) else None for i in META_TARGET_COLS]
            if all(v is None or (isinstance(v, str) and v.strip() == "") for v in picked):
                continue
            rows.append(picked)

        wb.close()

        df_f = pd.DataFrame(rows, columns=headers)
        df_f.insert(0, "RunName", runname)
        df_lst.append(df_f)

    return pd.concat(df_lst, ignore_index=True) if df_lst else pd.DataFrame()


# ══════════════════════════════════════════════════════════
# 4. SO KHỚP META ↔ SUM
# ══════════════════════════════════════════════════════════

def _norm(s: pd.Series) -> pd.Series:
    """Chuẩn hoá series để so khớp:
       - strip
       - bỏ phần '(...)' (ABC(XYZ) → ABC)
       - bỏ '.0' đuôi (1.0 → 1, fix mismatch int/float)
       - uppercase (case-insensitive)
       - 'NAN'/'NONE' → ''
    """
    return (
        s.astype(str)
         .str.strip()
         .str.replace(r"\s*\([^)]*\)", "", regex=True)
         .str.replace(r"\.0$", "", regex=True)
         .str.upper()
         .replace({"NAN": "", "NONE": ""})
    )


def add_sum_debug_keys(df_sum: pd.DataFrame) -> pd.DataFrame:
    """Thêm các cột key (đã normalize) vào df_sum để so trực tiếp với meta_only.xlsx.
       - _nipt_key  : RUN | C3 | order  (cho Sample Type='NIPT')
       - _crh_key   : RUN | C3          (cho Sample Type='CRH')
       - _other_key : RUN | C3          (cho Sample Type='other')
    """
    if df_sum.empty:
        return df_sum
    out = df_sum.copy()
    rn  = _norm(out["RunName"])
    c3  = _norm(out["Col 3"])
    odr = _norm(out["sample_order"])
    out["_nipt_key"]  = rn + "|" + c3 + "|" + odr   # type: ignore[operator]
    out["_crh_key"]   = rn + "|" + c3                # type: ignore[operator]
    out["_other_key"] = rn + "|" + c3                # type: ignore[operator]
    return out


def find_meta_only(df_meta: pd.DataFrame, df_sum: pd.DataFrame,
                   debug: bool = True) -> pd.DataFrame:
    """Trả về các dòng meta KHÔNG match trong df_sum.

    df_meta cột (theo thứ tự):
        RunName, expNum (col A), sampleOrder (col B), <U>, <V>
    df_sum cột:
        RunName, Col 3, Col 5, sample_order, Loại, Sample Type

    Match rule (1 trong 3):
      - NIPT  : SUM row Sample Type='NIPT' AND Loại='' AND
                RunName khớp AND Col 3 == expNum AND sample_order == sampleOrder
      - CRH   : SUM row Sample Type='CRH' AND Loại='' AND
                RunName khớp AND Col 3 == expNum
                (pool NIPT — KHÔNG cần sample_order)
      - other : SUM row Sample Type='other' AND
                RunName khớp AND Col 3 == (V nếu V≠'' else U)
                (KHÔNG cần sample_order)

    Output: 3 cột (RunName, expNum, sampleOrder) cho rows chưa match.
    Nếu debug=True: thêm _nipt_key_tried, _crh_key_tried, _other_key_tried.
    """
    cols = list(df_meta.columns)
    if len(cols) < 5:
        return df_meta.copy()

    rn_col, exp_col, ord_col, u_col, v_col = cols[:5]

    if df_meta.empty:
        out = df_meta[[rn_col, exp_col, ord_col]].copy()
        if debug:
            out["_nipt_key_tried"]  = ""
            out["_crh_key_tried"]   = ""
            out["_other_key_tried"] = ""
        return out

    # ── Build keys cho meta
    rn_m  = _norm(df_meta[rn_col])
    exp_m = _norm(df_meta[exp_col])
    ord_m = _norm(df_meta[ord_col])
    u_m   = _norm(df_meta[u_col])
    v_m   = _norm(df_meta[v_col])

    nipt_key  = rn_m + "|" + exp_m + "|" + ord_m   # type: ignore[operator]
    crh_key   = rn_m + "|" + exp_m                  # type: ignore[operator]
    other_id  = v_m.where(v_m.ne(""), u_m)          # V ưu tiên, fallback U
    other_key = rn_m + "|" + other_id               # type: ignore[operator]

    # ── Build set keys cho sum (lọc theo Sample Type)
    rn_s  = _norm(df_sum["RunName"])
    c3_s  = _norm(df_sum["Col 3"])
    ord_s = _norm(df_sum["sample_order"])

    nipt_sum_key  = rn_s + "|" + c3_s + "|" + ord_s   # type: ignore[operator]
    pool_sum_key  = rn_s + "|" + c3_s                  # type: ignore[operator]   (CRH + other dùng chung)

    loai_empty = df_sum["Loại"].astype(str).str.strip() == ""
    nipt_mask  = (df_sum["Sample Type"] == "NIPT") & loai_empty
    crh_mask   = (df_sum["Sample Type"] == "CRH")  & loai_empty
    other_mask = df_sum["Sample Type"] == "other"

    nipt_keys  = set(nipt_sum_key[nipt_mask])
    crh_keys   = set(pool_sum_key[crh_mask])
    other_keys = set(pool_sum_key[other_mask])

    matched = (
        nipt_key.isin(nipt_keys)
        | crh_key.isin(crh_keys)
        | other_key.isin(other_keys)
    )

    out = df_meta.loc[~matched, [rn_col, exp_col, ord_col]].copy()
    if debug:
        out["_nipt_key_tried"]  = nipt_key[~matched].values
        out["_crh_key_tried"]   = crh_key[~matched].values
        out["_other_key_tried"] = other_key[~matched].values
    return out.reset_index(drop=True)


# ══════════════════════════════════════════════════════════
# 5. STANDALONE — chọn folder, xuất file SUM đã xử lý
# ══════════════════════════════════════════════════════════
if __name__ == "__main__":
    import sys
    import tkinter as tk
    from tkinter import filedialog

    pd.set_option("display.max_columns", None)
    pd.set_option("display.width", 200)
    pd.set_option("display.max_rows", 200)

    root = tk.Tk()
    root.withdraw()
    folder = filedialog.askdirectory(title="Chọn folder chứa file SUM Excel")
    if not folder:
        print("Không chọn folder — thoát")
        sys.exit(0)

    candidates = sorted(glob.glob(os.path.join(folder, "SUM*.xlsx")))
    if not candidates:
        candidates = sorted(
            glob.glob(os.path.join(folder, "*.xlsx"))
            + glob.glob(os.path.join(folder, "*.xls"))
        )
    if not candidates:
        print(f"Không tìm thấy file Excel nào trong: {folder}")
        sys.exit(1)

    input_path = candidates[0]
    print(f"Đang xử lý: {input_path}")
    df_out = process_sum_file(input_path, sheet="Sheet1")
    print(df_out)
    print(f"\nShape: {df_out.shape}")

    in_stem  = os.path.splitext(os.path.basename(input_path))[0]
    out_path = os.path.join(folder, f"{in_stem}_solution.xlsx")
    df_out.to_excel(out_path, index=False)
    print(f"\n✓ Đã xuất: {out_path}")
