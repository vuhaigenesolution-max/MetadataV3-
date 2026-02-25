
# =====================
# Refactored for import
# =====================
import os
import pandas as pd
from openpyxl import load_workbook
import csv
from datetime import datetime

def process_excel(source_path, template_path, output_path):
    # 1. Đọc file nguồn và lấy cột T từ dòng 22
    wb_source = load_workbook(source_path, data_only=True)
    sheet_name = wb_source.sheetnames[0]
    ws = wb_source[sheet_name]
    t14_value = ws["T14"].value
    if t14_value is None or str(t14_value).strip() == "":
        raise ValueError("Ô T14 trong file nguồn đang trống, không thể tạo barcode header.")
    t14_str = str(t14_value).strip()
    parts = t14_str.split('_')
    run_name = None
    for p in parts:
        # Accept anything after the leading R (except empty), we sanitize later
        if p.upper().startswith('R') and len(p) > 1:
            run_name = p.upper()
            break
    if run_name is None:
        raise ValueError("Không tìm thấy Run name dạng Rxxx trong ô T14.")
    if len(parts) >= 2:
        run_date = f"{parts[-2]}_{parts[-1]}"
    else:
        run_date = t14_str
    # Clean filename parts to avoid invalid characters
    cleaned_run_name = "".join(c if c.isalnum() or c in ('-', '_') else '_' for c in run_name)
    safe_run_name = cleaned_run_name.upper()
    if not safe_run_name.startswith('R'):
        safe_run_name = f"R{safe_run_name}"
    if safe_run_name.strip('_') == "":
        raise ValueError("Run name không hợp lệ sau khi làm sạch (safe_run_name rỗng).")
    safe_run_date = "".join(c if c.isalnum() or c in ('-', '_') else '_' for c in run_date)
    barcode_name = t14_str
    safe_folder_name = "".join('_' if c in '<>:"/\\|?*' else c for c in t14_str).strip()
    if safe_folder_name == "":
        raise ValueError("Giá trị T14 không hợp lệ để đặt tên thư mục sau khi làm sạch.")
    values = []
    row = 22
    while True:
        value = ws[f"T{row}"].value
        if value is None:
            break
        values.append(value)
        row += 1
    df_T = pd.DataFrame(values, columns=["T"])

    # 2. Xóa dữ liệu cũ trong sheet 'MỒI T7' của template
    wb_template = load_workbook(template_path)
    ws_moi = wb_template["MỒI T7"]
    row = 2
    while ws_moi[f"A{row}"].value is not None:
        ws_moi[f"A{row}"].value = None
        row += 1
    wb_template.save(template_path)

    # 3. Tính toán VLOOKUP bằng Python
    wb = load_workbook(template_path, data_only=True)
    ws_primer = wb["PRIMER LOCKED"]
    rows_primer = list(ws_primer.iter_rows(values_only=True))
    max_col_primer = max(len(row) for row in rows_primer if row)
    rows_primer = [row[:max_col_primer] for row in rows_primer]
    df_primer = pd.DataFrame(rows_primer)
    if not df_primer.empty:
        df_primer.columns = [str(col) for col in df_primer.iloc[0]]
        df_primer = df_primer[1:].reset_index(drop=True)
    result_df = pd.DataFrame()
    result_df['A'] = df_T['T']
    lookup_dict_B = dict(zip(df_primer.iloc[:,0], df_primer.iloc[:,1]))
    result_df['B'] = result_df['A'].map(lookup_dict_B)
    lookup_dict_C = dict(zip(df_primer.iloc[:,1], df_primer.iloc[:,2]))
    result_df['C'] = result_df['B'].map(lookup_dict_C)
    df_bc = result_df[['B', 'C']].copy()
    wb.close()


    # 4. Tạo file CSV với 3 dòng đầu và dán result_df từ dòng 4
    rows = [
        ['#barcodeName', barcode_name],
        ['#misMatch1', 0],
        ['#misMatch2', 0]
    ]
    csv_filename = f'barcode.csv'
    target_folder = os.path.join(output_path, barcode_name)
    os.makedirs(target_folder, exist_ok=True)
    csv_path = os.path.join(target_folder, csv_filename)
    if os.path.exists(csv_path):
        raise FileExistsError(f"File đã tồn tại: {csv_path}. Vui lòng xóa hoặc đổi tên trước khi chạy lại.")
    with open(csv_path, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.writer(f)
        writer.writerows(rows)
        df_bc.to_csv(f, index=False, header=False)
    print(f"Đã tạo file CSV với 3 dòng đầu và dán result_df từ dòng 4 tại: {csv_path}")
    return csv_path


