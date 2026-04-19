"""
CombieFile.py
-------------
Gom và xuất dữ liệu từ nhiều file Excel nguồn vào file template.

Flow tổng quát:
  1. Quét folder → nhóm file theo (runname, date) từ tên file metadata_<run>_<date>.xlsx
  2. Đọc từng file (sheet "Sample", từ dòng 22), lọc dòng trống/không hợp lệ
  3. Concat các file cùng nhóm thành 1 DataFrame
  4. Kiểm tra ký tự đặc biệt / khoảng trắng → xuất file lỗi nếu có
  5. Copy template → paste data vào sheet "Sample" (cột A B C H I J T)
  6. Điền công thức cột K, L, M trong sheet "Sample" (VLOOKUP Index Sets)
  7. Điền công thức tất cả cột trong sheet "SampleImport" và "Aviti Manifest"
  8. Phát hiện collision trong SampleImport:
       - Đọc cột K, L từ df nguồn → tra cứu sequence trong sheet "Index Sequence"
       - So sánh từng cặp dòng: nếu diff(G_seq) < 3 VÀ diff(I_seq) < 3 → collision
       - Highlight màu vàng cột G, I các dòng liên quan
       - Ghi log tại Q24+: T | k_code | G_seq | l_code | I_seq (2 dòng/cặp + 1 dòng trống)
  9. Lưu file kết quả vào output_folder

Trả về: (output_files: list[str], total_collisions: int)
"""

import os
import re
import glob
import shutil
import string
import logging
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from collections import defaultdict

# ── Logger ────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)
# ──────────────────────────────────────────────────────────

# ── Cấu hình đọc file nguồn ───────────────────────────────
SHEET_SRC  = "Sample"
START_ROW  = 22          # dòng bắt đầu đọc & paste (Excel 1-indexed)

COL_IDX   = [0, 1, 2, 7, 8, 9, 10, 11, 19]    # vị trí cột trong file nguồn (0-indexed): A B C H I J K L T
COL_NAMES = ["A", "B", "C", "H", "I", "J", "K", "L", "T"]

# Mapping: tên cột df → số thứ tự cột trong sheet Sample của template (1-indexed)
COL_MAP = {
    "A": 1,   # cột A
    "B": 2,   # cột B
    "C": 3,   # cột C
    "H": 8,   # cột H
    "I": 9,   # cột I
    "J": 10,  # cột J
    "T": 20,  # cột T
}

FILE_PATTERN = re.compile(r"metadata_(.+?)_(\d{8}).*\.xlsx?$", re.IGNORECASE)

# ── Tên sheet tra cứu sequence ────────────────────────────
INDEX_SEQ_SHEET  = "Index Sequence"

# ── Collision detection ───────────────────────────────────
_COLLISION_FILL     = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
COLLISION_LOG_COL   = 17   # cột Q (1-indexed)
COLLISION_LOG_START = 23   # dòng 23 = header, dòng 24+ = data
# ──────────────────────────────────────────────────────────


# ══════════════════════════════════════════════════════════
# GHI VALUES TRỰC TIẾP — không dùng công thức Excel
# ══════════════════════════════════════════════════════════

IMPORT_SHEET_NAME = "SampleImport"
IMPORT_START_ROW  = 24

AVITI_SHEET_NAME = "Aviti Manifest"
AVITI_START_ROW  = 16


def _compute_test_type(sample_id, test_code) -> str:
    """Tính test type (cột K SampleImport) từ Sample_ID và test_code."""
    if pd.isna(sample_id):
        return ""
    if pd.isna(test_code):
        return str(sample_id).strip()
    a = str(sample_id).strip()
    c = str(test_code).strip()
    a1, a2, a4 = a[:1], a[:2], a[:4]
    c1, c2, c4 = c[:1], c[:2], c[:4]
    nipt_a = a1 in {"E", "H", "T", "B"} or a2 == "ID"
    if nipt_a and (c2 == "JI" or c1 == "I"):   return "TS1"
    if nipt_a and (c2 in {"JX", "JW"} or c1 == "X"): return "TS95"
    if nipt_a and (c2 == "JN" or c1 == "N"):   return "TS3"
    if nipt_a and c2 in {"JA", "AA", "JS", "SA"}: return "TS24"
    if a1 in {"T", "B"} and c2 == "AS":        return "TSPRO"
    if a1 in {"T", "B", "E", "H"} and c4 == "SERA": return "TSPRO"
    if a2 == "CR":                              return "CARRIER9"
    if a4 == "DEL3":                            return "NIPTDEL3"
    return a


def _complement(seq: str) -> str:
    """Complement chuỗi DNA: A↔T, C↔G."""
    return seq.translate(str.maketrans("ACGTacgt", "TGCAtgca"))


def _str(val) -> str:
    return "" if pd.isna(val) else str(val).strip()


def _fill_sample_values(ws, df, start_row):
    """Ghi values cột K, L vào sheet Sample; col M vẫn dùng formula PCR."""
    for i, (_, row) in enumerate(df.iterrows()):
        r = start_row + i
        ws.cell(row=r, column=11).value = _str(row.get("K"))
        ws.cell(row=r, column=12).value = _str(row.get("L"))
        ws.cell(row=r, column=13).value = (
            f"=VLOOKUP(A{r}&B{r},"
            f"CHOOSE({{1,2}},"
            f"PCR!$A$1:$A$3000&PCR!$B$1:$B$3000,"
            f"PCR!$C$1:$C$3000),2,0)"
        )
    log.info(f"  → Sheet '{ws.title}': ghi values K, L cho {len(df)} dòng (từ dòng {start_row})")


def _fill_import_values(ws, df, start_row, index_seq):
    """Ghi values trực tiếp vào sheet SampleImport."""
    for i, (_, row) in enumerate(df.iterrows()):
        r = start_row + i
        b, c_val = _str(row.get("B")), _str(row.get("C"))
        k_code   = _str(row.get("K"))
        l_code   = _str(row.get("L"))
        g_seq    = index_seq.get(k_code, "")
        i_seq    = index_seq.get(l_code, "")
        test_type = _compute_test_type(row.get("A"), row.get("C"))

        ws.cell(row=r, column=1).value  = f"{b}-{c_val}"   # A: Sample_ID
        ws.cell(row=r, column=2).value  = f"{b}-{c_val}"   # B: Sample_Name
        ws.cell(row=r, column=3).value  = _str(row.get("A"))# C: Sample barcode
        ws.cell(row=r, column=6).value  = k_code            # F: I7_Index_ID
        ws.cell(row=r, column=7).value  = g_seq             # G: index sequence
        ws.cell(row=r, column=8).value  = l_code            # H: I5_Index_ID
        ws.cell(row=r, column=9).value  = i_seq             # I: index2 sequence
        ws.cell(row=r, column=11).value = test_type         # K: test type
    log.info(f"  → Sheet '{ws.title}': ghi values cho {len(df)} dòng (từ dòng {start_row})")


def _fill_aviti_values(ws, df, start_row, index_seq):
    """Ghi values trực tiếp vào sheet Aviti Manifest."""
    for i, (_, row) in enumerate(df.iterrows()):
        r = start_row + i
        b, c_val = _str(row.get("B")), _str(row.get("C"))
        k_code   = _str(row.get("K"))
        l_code   = _str(row.get("L"))
        g_seq    = index_seq.get(k_code, "")
        i_seq    = index_seq.get(l_code, "")
        comp_i   = _complement(i_seq)
        test_type = _compute_test_type(row.get("A"), row.get("C"))

        ws.cell(row=r, column=1).value = f"{b}-{c_val}"    # A
        ws.cell(row=r, column=2).value = g_seq              # B: G sequence
        ws.cell(row=r, column=3).value = comp_i[::-1]       # C: reverse complement of I seq
        ws.cell(row=r, column=4).value = test_type          # D: test type
        ws.cell(row=r, column=9).value = comp_i             # I: complement of I seq
    log.info(f"  → Sheet '{ws.title}': ghi values cho {len(df)} dòng (từ dòng {start_row})")

# ══════════════════════════════════════════════════════════


# Cột cần kiểm tra ký tự đặc biệt / khoảng trắng / dấu câu
VALIDATE_COLS = ["A", "B", "C", "T"]

_PUNCTUATION_SET = set(string.punctuation)  # !"#$%&'()*+,-./:;<=>?@[\]^_`{|}~


def _check_cell(val):
    """Trả về list lỗi tìm thấy trong một giá trị ô."""
    s = str(val).strip()
    errors = []
    if re.search(r'\s', str(val)):          # khoảng trắng (bao gồm cả leading/trailing)
        errors.append("khoảng trắng")
    if any(c in _PUNCTUATION_SET for c in s):
        errors.append("dấu câu/ký tự đặc biệt")
    if any(ord(c) > 127 for c in s):       # ký tự ngoài ASCII
        errors.append("ký tự non-ASCII")
    return errors


def _validate_df(df, output_folder, runname, date):
    """
    Kiểm tra cột A,B,C,T trong df.
    Nếu có lỗi → xuất file error_<runname>_<date>.xlsx vào output_folder.
    Trả về đường dẫn file lỗi (hoặc None nếu không có lỗi).
    """
    error_rows = []
    for col in VALIDATE_COLS:
        if col not in df.columns:
            continue
        for i, val in enumerate(df[col]):
            if pd.isna(val):
                continue
            errs = _check_cell(val)
            if errs:
                error_rows.append({
                    "Cột":          col,
                    "Dòng Excel":   START_ROW + i,
                    "Giá trị":      val,
                    "Loại lỗi":     ", ".join(errs),
                })

    if not error_rows:
        return None

    df_err = pd.DataFrame(error_rows)
    err_name = f"error_{runname}_{date}.xlsx"
    err_path = os.path.join(output_folder, err_name)
    df_err.to_excel(err_path, index=False)
    log.warning(f"  ⚠ Phát hiện {len(error_rows)} ô lỗi → {err_name}")
    return err_path


def _read_source_file(fp):
    """Đọc 1 file nguồn, trả về df đã lọc hàng trống và chuyển ngày cột I."""
    fname = os.path.basename(fp)
    log.info(f"Đọc file: {fname}")
    df = pd.read_excel(
        fp,
        sheet_name=SHEET_SRC,
        header=None,
        skiprows=START_ROW - 1,
        usecols=COL_IDX,
    )
    df.columns = COL_NAMES
    before = len(df)

    # Loại hàng mà cột A (Sample ID) trống
    df = df.dropna(subset=["A"])

    # Loại hàng mà cả B lẫn C đều trống
    # (SampleImport cột A = B&"-"&C → sẽ chỉ còn "-" nếu cả hai trống)
    b_empty = df["B"].isna() | (df["B"].astype(str).str.strip() == "")
    c_empty = df["C"].isna() | (df["C"].astype(str).str.strip() == "")
    df = df[~(b_empty & c_empty)]

    # Chuyển ngày cột I
    df["I"] = pd.to_datetime(df["I"], errors="coerce").dt.strftime("%d/%m/%Y")

    log.info(f"  → {before} dòng đọc được, {len(df)} dòng hợp lệ (bỏ {before - len(df)} hàng trống)")
    return df


def _count_char_diff(s1: str, s2: str) -> int:
    """Đếm số vị trí ký tự khác nhau giữa 2 chuỗi (ký tự thừa tính là khác)."""
    max_len = max(len(s1), len(s2))
    diffs = 0
    for i in range(max_len):
        c1 = s1[i] if i < len(s1) else ""
        c2 = s2[i] if i < len(s2) else ""
        if c1 != c2:
            diffs += 1
    return diffs


def _build_index_seq(wb):
    """Đọc sheet 'Index Sequence': code → sequence_str."""
    index_seq = {}
    if INDEX_SEQ_SHEET in wb.sheetnames:
        for row in wb[INDEX_SEQ_SHEET].iter_rows(min_row=2, values_only=True):
            if row[0] is not None and len(row) > 1:
                index_seq[str(row[0]).strip()] = str(row[1]).strip() if row[1] is not None else ""
    else:
        log.warning(f"Sheet '{INDEX_SEQ_SHEET}' không tìm thấy — bỏ qua collision detection")
    return index_seq


def _detect_and_mark_collisions(wb, df, import_start_row: int) -> int:
    """
    Phát hiện collision trong SampleImport bằng cách tra cứu sequence trực tiếp
    từ các sheet 'Index Sets' và 'Index Sequence' trong template (không cần Excel tính formula).

    Collision khi: diff(G_seq) < 3 VÀ diff(I_seq) < 3.

    Kết quả:
      - Highlight màu vàng cột G (7) và I (9) của các dòng collision.
      - Ghi log tại Q23 (header) và Q24+ (data: F, G, H, I của từng cặp).
    """
    if IMPORT_SHEET_NAME not in wb.sheetnames:
        return 0

    index_seq = _build_index_seq(wb)
    if not index_seq:
        return 0

    # Lấy k_code/l_code trực tiếp từ cột K và L của df (không qua Index Sets)
    rows_info = []
    for _, row in df.iterrows():
        t_val = str(row.get("T", "")).strip() if not pd.isna(row.get("T", pd.NA)) else ""
        k_code = str(row.get("K", "")).strip() if not pd.isna(row.get("K", pd.NA)) else ""
        l_code = str(row.get("L", "")).strip() if not pd.isna(row.get("L", pd.NA)) else ""
        rows_info.append({
            "t_val":  t_val,
            "k_code": k_code,
            "l_code": l_code,
            "g_seq":  index_seq.get(k_code, ""),
            "i_seq":  index_seq.get(l_code, ""),
        })

    # Duyệt tất cả cặp để tìm collision
    n = len(rows_info)
    collisions = []
    for a in range(n):
        for b in range(a + 1, n):
            if (_count_char_diff(rows_info[a]["g_seq"], rows_info[b]["g_seq"]) < 3 and
                    _count_char_diff(rows_info[a]["i_seq"], rows_info[b]["i_seq"]) < 3):
                collisions.append((a, b))

    if not collisions:
        log.info("  Không phát hiện collision trong SampleImport")
        return 0

    log.warning(f"  ⚠ Phát hiện {len(collisions)} cặp collision trong SampleImport")

    ws = wb[IMPORT_SHEET_NAME]

    # ── Highlight cột G (7) và I (9) ─────────────────────
    highlighted = set()
    for a, b in collisions:
        highlighted.add(a)
        highlighted.add(b)
    for row_idx in highlighted:
        r = import_start_row + row_idx
        ws.cell(row=r, column=7).fill = _COLLISION_FILL
        ws.cell(row=r, column=9).fill = _COLLISION_FILL

    # ── Ghi log vào Q24+ — mỗi cặp: 2 dòng + 1 dòng trống ──
    # Bố cục: F | G | H | I  (cột Q, R, S, T)
    #   dòng 1: dữ liệu row A
    #   dòng 2: dữ liệu row B
    #   dòng 3: trống (phân cách)
    write_row = COLLISION_LOG_START + 1   # bắt đầu từ dòng 24

    for a, b in collisions:
        ri_a, ri_b = rows_info[a], rows_info[b]

        for j, v in enumerate([ri_a["t_val"], ri_a["k_code"], ri_a["g_seq"], ri_a["l_code"], ri_a["i_seq"]]):
            ws.cell(row=write_row,     column=COLLISION_LOG_COL + j).value = v
        for j, v in enumerate([ri_b["t_val"], ri_b["k_code"], ri_b["g_seq"], ri_b["l_code"], ri_b["i_seq"]]):
            ws.cell(row=write_row + 1, column=COLLISION_LOG_COL + j).value = v

        write_row += 3   # 2 dòng data + 1 dòng trống
        log.warning(f"    Collision: Excel dòng {import_start_row + a} ↔ dòng {import_start_row + b}")

    return len(collisions)


def _paste_to_template(df, template_path, out_path, sheet_name):
    """Copy template → paste data vào sheet Sample → kéo công thức cả 2 sheet."""
    log.info(f"Copy template → {os.path.basename(out_path)}")
    shutil.copy2(template_path, out_path)
    wb = load_workbook(out_path)
    num_rows = len(df)

    # ── Sheet Sample: paste data ──────────────────────────
    ws_sample = wb[sheet_name]

    for row_idx, (_, row_data) in enumerate(df.iterrows()):
        excel_row = START_ROW + row_idx
        for df_col, col_num in COL_MAP.items():
            if df_col not in row_data:
                continue
            ws_sample.cell(row=excel_row, column=col_num).value = row_data[df_col]

    log.info(f"  → Đã paste {num_rows} dòng data vào sheet '{ws_sample.title}'")

    # Build index_seq một lần, dùng chung cho cả 3 sheet
    index_seq = _build_index_seq(wb)

    # ── Sheet Sample: ghi values cột K, L, M ─────────────
    _fill_sample_values(ws_sample, df, START_ROW)

    # ── Sheet SampleImport: ghi values ──────────────────
    if IMPORT_SHEET_NAME in wb.sheetnames:
        ws_import = wb[IMPORT_SHEET_NAME]
        _fill_import_values(ws_import, df, IMPORT_START_ROW, index_seq)

        # Xóa các dòng thừa trong template
        first_extra = IMPORT_START_ROW + num_rows
        if ws_import.max_row >= first_extra:
            for r in range(first_extra, ws_import.max_row + 1):
                for col in range(1, ws_import.max_column + 1):
                    ws_import.cell(row=r, column=col).value = None
            log.info(f"  → Xóa dòng thừa SampleImport: dòng {first_extra}–{ws_import.max_row}")
    else:
        log.warning(f"Không tìm thấy sheet '{IMPORT_SHEET_NAME}' trong template — bỏ qua")

    # ── Sheet SampleImport: phát hiện & đánh dấu collision ─
    collision_count = _detect_and_mark_collisions(wb, df, IMPORT_START_ROW)

    # ── Sheet Aviti Manifest: ghi values ─────────────────
    if AVITI_SHEET_NAME in wb.sheetnames:
        ws_aviti = wb[AVITI_SHEET_NAME]
        _fill_aviti_values(ws_aviti, df, AVITI_START_ROW, index_seq)
    else:
        log.warning(f"Không tìm thấy sheet '{AVITI_SHEET_NAME}' trong template — bỏ qua")

    wb.save(out_path)
    wb.close()
    log.info(f"  → Lưu thành công: {os.path.basename(out_path)}")
    return collision_count


def process_combine_files(folder, template_path, output_folder, sheet_template="Sample",
                          progress_callback=None, filter_file=None):
    """
    Duyệt tất cả file Excel trong folder, gom nhóm theo (runname, date),
    concat dữ liệu từng nhóm rồi paste vào template, xuất file mới vào output_folder.
    """
    os.makedirs(output_folder, exist_ok=True)

    files = (
        glob.glob(os.path.join(folder, "*.xlsx"))
        + glob.glob(os.path.join(folder, "*.xls"))
    )
    if filter_file:
        files = [f for f in files if os.path.basename(f) == filter_file]
        log.info(f"Tìm thấy {len(files)} file Excel (đã lọc: {filter_file})")
    else:
        log.info(f"Tìm thấy {len(files)} file Excel trong: {folder}")
    if not files:
        raise FileNotFoundError(f"Không tìm thấy file Excel nào trong: {folder}")

    groups = defaultdict(list)
    skipped = []
    for fp in files:
        m = FILE_PATTERN.search(os.path.basename(fp))
        if m:
            groups[(m.group(1), m.group(2))].append(fp)
            log.info(f"  ✔ {os.path.basename(fp)}  →  run={m.group(1)}, date={m.group(2)}")
        else:
            skipped.append(os.path.basename(fp))
            log.warning(f"  ✘ {os.path.basename(fp)}  →  không khớp định dạng, bỏ qua")

    log.info(f"Gom được {len(groups)} nhóm, bỏ qua {len(skipped)} file")

    total = len(groups)
    output_files = []
    collision_per_file: dict[str, int] = {}
    error_files: list[str] = []

    for idx, ((runname, date), fps) in enumerate(groups.items(), start=1):
        log.info(f"--- Nhóm {idx}/{total}: run={runname}, date={date} ({len(fps)} file) ---")
        dfs = []
        for fp in fps:
            try:
                dfs.append(_read_source_file(fp))
            except Exception as e:
                log.error(f"Đọc {os.path.basename(fp)}: {e}")

        if not dfs:
            log.warning(f"({runname}, {date}): không có dữ liệu hợp lệ, bỏ qua.")
            continue

        df_merged = pd.concat(dfs, ignore_index=True)
        log.info(f"Tổng sau concat: {len(df_merged)} dòng")

        # ── Kiểm tra ký tự đặc biệt / khoảng trắng / dấu câu ──
        err_path = _validate_df(df_merged, output_folder, runname, date)
        if err_path:
            error_files.append(os.path.basename(err_path))

        out_name = f"metadata_{runname}_{date}.xlsx"
        out_path = os.path.join(output_folder, out_name)

        if os.path.exists(out_path):
            log.warning(f"File đã tồn tại, sẽ ghi đè: {out_name}")

        col_count = _paste_to_template(df_merged, template_path, out_path, sheet_template)
        collision_per_file[out_name] = col_count
        output_files.append(out_path)

        if progress_callback:
            progress_callback(idx, total)

    total_collisions = sum(collision_per_file.values())
    log.info(f"Hoàn tất! Xuất {len(output_files)}/{total} file → {output_folder}")
    if total_collisions:
        log.warning(f"Tổng collision phát hiện: {total_collisions} cặp")
    return output_files, {
        "total_collisions": total_collisions,
        "collision_per_file": collision_per_file,
        "error_files": error_files,
    }


