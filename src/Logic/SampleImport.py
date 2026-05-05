"""
SampleImport.py
---------------
Xử lý file metadata Excel đã được combine, tạo 2 file CSV output:
  Output 1 : SampleImport_<tên file>.csv   — từ sheet SampleImport
  Output 2 : <tên file>_<date>.csv         — từ sheet Aviti Manifest

Luồng chính:
  1. Đọc file nguồn (phải có sheet "Sample")
  2. Xác định dòng cuối dữ liệu theo cột A của sheet Sample
  3. Đọc values-only (không formula) từ SampleImport và Aviti Manifest
  4. Làm sạch: replace 0 → ""
  5. Chuẩn hoá test-type cột K (tsprocare→TSPRO, tsthalass→TS3, ts9.5→TS95, còn lại→UPPERCASE)
  6. Auto-detect vùng từ runname (R→Nam, P→Bắc) → kiểm tra theo vùng:
       Vùng Nam : cột J không được trống nếu kit là T7/MGI (tra từ Nhật ký dò miền Nam)
       Vùng Bắc : cột Sample Project không được trống nếu Sequencer chứa G99 (tra từ Nhật ký dò miền Bắc)
       → Trả về DataFrame j_errors (File, Runname, Kit/Sequencer, SampleID, ColJ/SampleProject)
  7. Kiểm tra cột Description so với bảng gói xét nghiệm (goi_xn_path)
       → Trả về DataFrame desc_errors (File, SampleID, SamplePlate, Description, Description check)
  8. Xuất 2 file CSV
  ※ Việc gom và xuất file cảnh báo Excel (warning_col_J_SampleProject.xlsx,
     warning_description.xlsx) được thực hiện ở tầng backend sau khi xử lý tất cả file.
"""

import os
import re
import logging
import pandas as pd
from openpyxl import load_workbook
from Logic.CheckDescription import load_labcode_lookup

log = logging.getLogger(__name__)

# ── Tên sheet ─────────────────────────────────────────────
SHEET_SAMPLE  = "Sample"
SHEET_IMPORT  = "SampleImport"
SHEET_AVITI   = "Aviti Manifest"

# ── Dòng bắt đầu dữ liệu (1-indexed, Excel) ──────────────
SAMPLE_START_ROW  = 22
IMPORT_START_ROW  = 24
IMPORT_HEADER_ROW = 23   # dòng header ngay trên dữ liệu SampleImport
AVITI_START_ROW   = 16

# ── Cột A (0-indexed) dùng xác định dòng cuối ─────────────
COL_A = 0

# ── Mapping test-type (lowercase key) ─────────────────────
TEST_TYPE_MAP = {
    "tsprocare": "TSPRO",
    "tsthalass": "TS3",
    "ts9.5":     "TS95",
}

# ── Regex lấy runname + date từ tên file ──────────────────
FILE_PATTERN = re.compile(r"metadata_(.+?)_(\d{8})", re.IGNORECASE)


# ══════════════════════════════════════════════════════════
# HÀM HỖ TRỢ
# ══════════════════════════════════════════════════════════

def _get_last_data_row(ws, start_row: int) -> int:
    """
    Trả về index dòng cuối (1-indexed) có giá trị khác None / "" ở cột A.
    Duyệt từ start_row xuống cuối sheet.
    """
    last = start_row - 1
    for row in ws.iter_rows(min_row=start_row, min_col=1, max_col=1):
        cell_val = row[0].value
        if cell_val is not None and str(cell_val).strip() != "":
            last = row[0].row
    return last


def _sheet_to_df(wb, sheet_name: str, start_row: int, last_row: int) -> pd.DataFrame:
    """
    Đọc dữ liệu values-only (không formula) từ một sheet của workbook.
    Trả về DataFrame, cột đặt tên theo chữ cái Excel (A, B, C, ...).
    """
    if sheet_name not in wb.sheetnames:
        log.warning(f"Sheet '{sheet_name}' không tồn tại trong file — trả về DataFrame rỗng")
        return pd.DataFrame()

    ws = wb[sheet_name]
    max_col = ws.max_column

    rows = []
    for row in ws.iter_rows(min_row=start_row, max_row=last_row,
                             min_col=1, max_col=max_col):
        rows.append([cell.value for cell in row])

    # Đặt tên cột theo chữ cái Excel: A, B, C, ... Z, AA, AB, ...
    col_names = _excel_col_names(max_col)
    df = pd.DataFrame(rows, columns=col_names)
    return df


def _excel_col_names(n: int) -> list[str]:
    """Sinh tên cột Excel: A, B, ..., Z, AA, AB, ... cho n cột."""
    names = []
    for i in range(n):
        name, col = "", i + 1
        while col:
            col, rem = divmod(col - 1, 26)
            name = chr(65 + rem) + name
        names.append(name)
    return names


def _read_header_row(wb, sheet_name: str, header_row: int) -> list[str]:
    """Đọc dòng header → list tên cột thực tế (giữ thứ tự, ô trống = '')."""
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    headers = []
    for cell in ws[header_row]:
        v = cell.value
        headers.append("" if v is None else str(v).strip())
    return headers


def _rename_with_headers(df: pd.DataFrame, headers: list[str]) -> pd.DataFrame:
    """Đổi tên cột theo position: A→headers[0], B→headers[1]... Header rỗng giữ nguyên letter."""
    if df.empty or not headers:
        return df
    new_names = []
    for i, old in enumerate(df.columns):
        if i < len(headers) and headers[i]:
            new_names.append(headers[i])
        else:
            new_names.append(old)
    df = df.copy()
    df.columns = new_names
    return df


def _clean_zeros(df: pd.DataFrame) -> pd.DataFrame:
    """Thay thế các ô có giá trị 0 (hoặc "0") thành chuỗi rỗng ""."""
    return df.replace({0: "", "0": ""})


def _extract_runname(filename: str) -> str:
    """metadata_R9827_20260413.xlsx → 'R9827'"""
    m = FILE_PATTERN.search(filename)
    return m.group(1) if m else ""


def _check_col_j_empty(df_import: pd.DataFrame, fname: str, runname: str,
                        nhat_ky_path: str) -> pd.DataFrame:
    """
    Nếu runname khớp cột A nhật ký dò VÀ cột C (kit chạy) chứa 'T7' hoặc 'MGI',
    kiểm tra cột J của df_import (SampleImport từ dòng 24) không được trống.
    Trả về DataFrame các dòng lỗi, hoặc DataFrame rỗng nếu không có vấn đề.
    """
    empty_df = pd.DataFrame()

    if not nhat_ky_path or not os.path.isfile(nhat_ky_path):
        return empty_df
    if not runname:
        return empty_df

    try:
        df_nk = pd.read_excel(nhat_ky_path, header=0)
    except Exception as e:
        log.warning(f"Không đọc được nhật ký dò: {e}")
        return empty_df

    if df_nk.shape[1] < 3:
        return empty_df

    col_a = df_nk.columns[0]
    col_c = df_nk.columns[2]

    matched = df_nk[df_nk[col_a].astype(str).str.strip().str.upper() == runname.upper()]
    if matched.empty:
        log.info(f"  Runname '{runname}' không tìm thấy trong nhật ký dò — bỏ qua kiểm tra cột J")
        return empty_df

    kit_chay = str(matched.iloc[0][col_c]).upper()
    if "T7" not in kit_chay and "MGI" not in kit_chay:
        log.info(f"  Kit '{kit_chay}' không phải T7/MGI — bỏ qua kiểm tra cột J")
        return empty_df

    if "J" not in df_import.columns:
        log.warning(f"  Không có cột J trong SampleImport của '{fname}'")
        return empty_df

    is_empty = df_import["J"].isna() | (df_import["J"].astype(str).str.strip().isin(["", "None", "nan"]))
    df_err = df_import[is_empty].copy()

    if df_err.empty:
        return empty_df

    result = pd.DataFrame({
        "File":    fname,
        "Runname": runname,
        "Kit":     kit_chay,
        "SampleID": df_err["A"] if "A" in df_err.columns else "",
        "ColJ":    df_err["J"],
    })
    return result.reset_index(drop=True)


def _col_letter_from_idx(col_idx: int) -> str:
    """1-based column index → Excel letter (1→'A', 26→'Z', 27→'AA', ...)."""
    name, col = "", col_idx
    while col:
        col, rem = divmod(col - 1, 26)
        name = chr(65 + rem) + name
    return name


def _find_col_letter(ws, header_row: int, col_name: str) -> str:
    """Tìm tên cột Excel (A/B/AA…) ứng với header có tên col_name trong dòng header_row."""
    for cell in ws[header_row]:
        if cell.value and str(cell.value).strip().lower() == col_name.strip().lower():
            return _col_letter_from_idx(cell.column)
    return ""


def _check_sample_project_empty(df_import: pd.DataFrame, fname: str, runname: str,
                                 nhat_ky_path: str, sample_project_col: str) -> pd.DataFrame:
    """
    Vùng Bắc: nếu runname khớp cột A nhật ký dò VÀ cột D (Sequencer) chứa 'G99',
    kiểm tra cột Sample Project của df_import (SampleImport từ dòng 24) không được trống.
    Trả về DataFrame các dòng lỗi, hoặc DataFrame rỗng nếu không có vấn đề.
    """
    empty_df = pd.DataFrame()

    if not nhat_ky_path or not os.path.isfile(nhat_ky_path):
        return empty_df
    if not runname or not sample_project_col:
        return empty_df

    try:
        df_nk = pd.read_excel(nhat_ky_path, header=0)
    except Exception as e:
        log.warning(f"Không đọc được nhật ký dò: {e}")
        return empty_df

    if df_nk.shape[1] < 4:
        return empty_df

    col_a = df_nk.columns[0]   # Đợt chạy
    col_d = df_nk.columns[3]   # Sequencer

    matched = df_nk[df_nk[col_a].astype(str).str.strip().str.upper() == runname.upper()]
    if matched.empty:
        log.info(f"  Runname '{runname}' không tìm thấy trong nhật ký dò — bỏ qua kiểm tra Sample Project")
        return empty_df

    sequencer = str(matched.iloc[0][col_d]).upper()
    if "G99" not in sequencer:
        log.info(f"  Sequencer '{sequencer}' không chứa G99 — bỏ qua kiểm tra Sample Project")
        return empty_df

    if sample_project_col not in df_import.columns:
        log.warning(f"  Không tìm thấy cột '{sample_project_col}' (Sample Project) trong '{fname}'")
        return empty_df

    is_empty = (
        df_import[sample_project_col].isna()
        | df_import[sample_project_col].astype(str).str.strip().isin(["", "None", "nan"])
    )
    df_err = df_import[is_empty].copy()

    if df_err.empty:
        return empty_df

    result = pd.DataFrame({
        "File":           fname,
        "Runname":        runname,
        "Sequencer":      sequencer,
        "SampleID":       df_err["A"] if "A" in df_err.columns else "",
        "SampleProject":  df_err[sample_project_col],
    })
    return result.reset_index(drop=True)


def _check_description_mismatch(source_path: str, fname: str, goi_xn_path: str) -> pd.DataFrame:
    """
    Đọc sheet SampleImport (dòng header 23, data từ 24) bằng pandas,
    so sánh cột Description thực tế với giá trị tra từ bảng labcode (goi_xn_path).
    Trả về DataFrame các dòng lệch, hoặc DataFrame rỗng nếu không có vấn đề.
    """
    if not goi_xn_path or not os.path.isfile(goi_xn_path):
        return pd.DataFrame()

    try:
        lookup = load_labcode_lookup(goi_xn_path)

        # skiprows = IMPORT_HEADER_ROW - 1 = 22 → dòng 23 thành header
        df = pd.read_excel(source_path, sheet_name=SHEET_IMPORT,
                           skiprows=IMPORT_HEADER_ROW - 1, header=0)
        df = df.dropna(axis=1, how="all").dropna(axis=0, how="all").reset_index(drop=True)

        if "Description" not in df.columns:
            return pd.DataFrame()

        col_sample_id    = df.columns[0]
        col_sample_plate = df.columns[2]

        def _pfx(val, n=3):
            if pd.isna(val): return ""
            return re.sub(r'[\d\-]', '', str(val).strip())[:n].upper()

        def _nhom(prefix):
            if prefix[:2] == "GS": return "SGNU"
            if prefix[:2] == "CR": return "CRH"
            if prefix[:1] in {"E","T","V","H","B","L","P"}: return "NIPT"
            return ""

        df["_prefix"] = df[col_sample_id].apply(lambda x: _pfx(x, 3))
        df["_nhom"]   = df[col_sample_plate].apply(lambda x: _nhom(_pfx(x, 3)))

        def _expected(row):
            nhom = row["_nhom"]
            if nhom == "SGNU": return "SGNU"
            result = lookup.get((row["_prefix"], nhom), "")
            return result if result else (str(row["Description"]).strip() if not pd.isna(row["Description"]) else "")

        df["_desc_check"] = df.apply(_expected, axis=1)

        mask   = df["Description"].astype(str).str.strip() != df["_desc_check"].astype(str).str.strip()
        df_err = df[mask][[col_sample_id, col_sample_plate, "Description", "_desc_check"]].copy()
        df_err = df_err.rename(columns={"_desc_check": "Description check"})

        if df_err.empty:
            return pd.DataFrame()

        df_err.insert(0, "File", fname)
        return df_err.reset_index(drop=True)

    except Exception as e:
        log.warning(f"  Lỗi kiểm tra Description: {e}")
        return pd.DataFrame()


def _map_test_type(val) -> str:
    """Chuẩn hoá test-type: tsprocare→TSPRO, tsthalass→TS3, ts9.5→TS95, còn lại→UPPERCASE."""
    if pd.isna(val) or str(val).strip() == "":
        return val
    key = str(val).strip().lower()
    return TEST_TYPE_MAP.get(key, str(val).strip().upper())



# ══════════════════════════════════════════════════════════
# HÀM CHÍNH
# ══════════════════════════════════════════════════════════

def process_sample_import(source_path: str,
                           output_folder: str,
                           nhat_ky_nam_path: str = "",
                           nhat_ky_bac_path: str = "",
                           goi_xn_path: str = "") -> dict:
    """
    Xử lý một file metadata Excel đã combine.

    Vùng được tự động phát hiện từ tên file:
      - Runname bắt đầu bằng R (Rxxx) → miền Nam → dùng nhat_ky_nam_path
      - Runname bắt đầu bằng P (Pxxx) → miền Bắc → dùng nhat_ky_bac_path

    Args:
        source_path      : đường dẫn file Excel đầu vào
        output_folder    : thư mục lưu 2 file CSV output
        nhat_ky_nam_path : đường dẫn file Nhật ký dò miền Nam (tuỳ chọn)
        nhat_ky_bac_path : đường dẫn file Nhật ký dò miền Bắc (tuỳ chọn)
        goi_xn_path      : đường dẫn file Thông tin gói xét nghiệm (tuỳ chọn)

    Returns:
        dict với keys "csv_import" và "csv_aviti" là đường dẫn file đã xuất.
    """
    os.makedirs(output_folder, exist_ok=True)

    fname   = os.path.basename(source_path)
    runname = _extract_runname(fname)

    m = FILE_PATTERN.search(fname)
    date_str = m.group(2) if m else "00000000"

    # ── Auto-detect vùng từ runname ───────────────────────
    _re_nam = re.compile(r"^R\d+", re.IGNORECASE)
    _re_bac = re.compile(r"^P\d+", re.IGNORECASE)
    if _re_nam.match(runname):
        vung = "Nam"
        nhat_ky_path = nhat_ky_nam_path
    elif _re_bac.match(runname):
        vung = "Bắc"
        nhat_ky_path = nhat_ky_bac_path
    else:
        vung = ""
        nhat_ky_path = ""
        log.warning(f"  Không xác định được vùng từ runname '{runname}' — bỏ qua kiểm tra nhật ký dò")

    log.info(f"Xử lý file: {fname} | Runname: {runname} | Vùng: {vung or 'không xác định'}")
    if nhat_ky_path:
        log.info(f"  Nhật ký dò  : {nhat_ky_path}")
    if goi_xn_path:
        log.info(f"  Gói xét nghiệm: {goi_xn_path}")

    # ── 1. Mở workbook (data_only=True để lấy values thay vì formula) ──
    try:
        wb = load_workbook(source_path, data_only=True, read_only=True)
    except Exception as e:
        raise ValueError(f"Không mở được file Excel: {e}")

    # ── 2. Kiểm tra sheet Sample ──────────────────────────
    if SHEET_SAMPLE not in wb.sheetnames:
        wb.close()
        raise ValueError(f"File '{fname}' không có sheet '{SHEET_SAMPLE}' — bỏ qua")

    ws_sample = wb[SHEET_SAMPLE]

    # ── 3. Xác định dòng cuối theo cột A của sheet Sample ─
    last_row_sample = _get_last_data_row(ws_sample, SAMPLE_START_ROW)
    num_rows        = last_row_sample - SAMPLE_START_ROW + 1

    if num_rows <= 0:
        wb.close()
        log.warning(f"Sheet Sample không có dữ liệu (file: {fname})")
        return {}

    log.info(f"  Dòng dữ liệu Sample: {SAMPLE_START_ROW} → {last_row_sample} ({num_rows} dòng)")

    # Tính dòng cuối tương ứng cho các sheet khác (cùng số dòng dữ liệu)
    last_row_import = IMPORT_START_ROW + num_rows - 1
    last_row_aviti  = AVITI_START_ROW  + num_rows - 1

    # ── 4. Đọc SampleImport ───────────────────────────────
    df_import = _sheet_to_df(wb, SHEET_IMPORT, IMPORT_START_ROW, last_row_import)
    log.info(f"  SampleImport: {len(df_import)} dòng × {len(df_import.columns)} cột")

    # ── 5. Đọc Aviti Manifest ─────────────────────────────
    df_aviti = _sheet_to_df(wb, SHEET_AVITI, AVITI_START_ROW, last_row_aviti)
    log.info(f"  Aviti Manifest: {len(df_aviti)} dòng × {len(df_aviti.columns)} cột")

    # ── 5b. Tìm cột "Sample Project" cho kiểm tra vùng Bắc ──
    sample_project_col = ""
    if vung == "Bắc" and SHEET_IMPORT in wb.sheetnames:
        sample_project_col = _find_col_letter(wb[SHEET_IMPORT], IMPORT_HEADER_ROW, "Sample Project")
        if sample_project_col:
            log.info(f"  Cột 'Sample Project' = {sample_project_col}")
        else:
            log.warning("  Không tìm thấy cột 'Sample Project' trong header SampleImport")

    # nhat_ky_path đã được gán khi auto-detect vùng ở trên

    # ── 5c. Đọc header thực để xuất CSV với tên cột đúng ──
    import_headers = _read_header_row(wb, SHEET_IMPORT, IMPORT_HEADER_ROW)
    aviti_headers  = _read_header_row(wb, SHEET_AVITI,  AVITI_START_ROW - 1)

    wb.close()

    # ── 6. Làm sạch: 0 → "" ──────────────────────────────
    df_import = _clean_zeros(df_import)
    df_aviti  = _clean_zeros(df_aviti)

    # ── 7. Chuẩn hoá test-type cột K ─────────────────────────
    if not df_import.empty and "K" in df_import.columns:
        df_import = df_import.copy()
        df_import["K"] = df_import["K"].apply(_map_test_type)

    # ── 8. Kiểm tra theo vùng (auto-detected) ────────────
    if vung == "Nam":
        df_j_errors = _check_col_j_empty(df_import, fname, runname, nhat_ky_path)
        if not df_j_errors.empty:
            log.warning(f"  ⚠ Cột J trống ({len(df_j_errors)} dòng) — kit={df_j_errors['Kit'].iloc[0]}")
    elif vung == "Bắc":
        df_j_errors = _check_sample_project_empty(df_import, fname, runname, nhat_ky_path, sample_project_col)
        if not df_j_errors.empty:
            log.warning(f"  ⚠ Sample Project trống ({len(df_j_errors)} dòng) — sequencer={df_j_errors['Sequencer'].iloc[0]}")
    else:
        df_j_errors = pd.DataFrame()

    # ── 9. Kiểm tra Description so với bảng labcode ─────────
    df_desc_errors = _check_description_mismatch(source_path, fname, goi_xn_path)
    if not df_desc_errors.empty:
        log.warning(f"  ⚠ Description lệch ({len(df_desc_errors)} dòng)")

    # ── 10. Xuất Output 1: SampleImport_<runname>.csv ─────
    csv1_name = f"SampleImport_{runname}.csv"
    csv1_path = os.path.join(output_folder, csv1_name)
    _rename_with_headers(df_import, import_headers).to_csv(
        csv1_path, index=False, encoding="utf-8-sig"
    )
    log.info(f"  → Output 1: {csv1_name}")

    # ── 11. Xuất Output 2: Manifest_<runname>_<yyyymmdd>.csv ──
    csv2_name = f"Manifest_{runname}_{date_str}.csv"
    csv2_path = os.path.join(output_folder, csv2_name)
    _rename_with_headers(df_aviti, aviti_headers).to_csv(
        csv2_path, index=False, encoding="utf-8-sig"
    )
    log.info(f"  → Output 2: {csv2_name}")

    return {
        "csv_import":     csv1_path,
        "csv_aviti":      csv2_path,
        "j_errors":       df_j_errors,
        "desc_errors":    df_desc_errors,
    }
